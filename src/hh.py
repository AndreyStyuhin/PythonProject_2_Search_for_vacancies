import abc
import requests
import json
import csv
import openpyxl
from typing import List, Dict, Any, Optional


class VacancyAPI(abc.ABC):
    """Абстрактный класс для работы с API сервисов с вакансиями."""

    @abc.abstractmethod
    def get_vacancies(self, search_query: str) -> List[Dict[str, Any]]:
        pass


class HHVacancyAPI(VacancyAPI):
    """Класс для работы с API hh.ru."""

    def __init__(self):
        self.base_url = "https://api.hh.ru/vacancies"

    def get_vacancies(self, search_query: str) -> List[Dict[str, Any]]:
        params = {
            "text": search_query,
            "area": 113,  # Россия
            "per_page": 100,  # Количество вакансий на странице
        }
        response = requests.get(self.base_url, params=params)
        response.raise_for_status()
        return response.json().get("items", [])


class Vacancy:
    """Класс для представления вакансии."""

    def __init__(
        self,
        title: str,
        link: str,
        salary: Optional[Dict[str, Optional[int]]],
        description: str,
        requirements: str,
    ):
        self.title = title
        self.link = link
        self.salary = salary
        self.description = description
        self.requirements = requirements

    def __repr__(self):
        return (
            f"Vacancy(title='{self.title}', link='{self.link}', "
            f"salary={self.salary}, description='{self.description[:50]}...', "
            f"requirements='{self.requirements[:50]}...')"
        )

    def __eq__(self, other):
        if not isinstance(other, Vacancy):
            return False
        return self.get_salary() == other.get_salary()

    def __lt__(self, other):
        if not isinstance(other, Vacancy):
            return NotImplemented
        return self.get_salary() < other.get_salary()

    def __le__(self, other):
        if not isinstance(other, Vacancy):
            return NotImplemented
        return self.get_salary() <= other.get_salary()

    def __gt__(self, other):
        if not isinstance(other, Vacancy):
            return NotImplemented
        return self.get_salary() > other.get_salary()

    def __ge__(self, other):
        if not isinstance(other, Vacancy):
            return NotImplemented
        return self.get_salary() >= other.get_salary()

    def get_salary(self) -> int:
        """Возвращает среднюю зарплату или 0, если зарплата не указана."""
        if not self.salary:
            return 0
        from_salary = self.salary.get("from")
        to_salary = self.salary.get("to")
        if from_salary and to_salary:
            return (from_salary + to_salary) // 2
        elif from_salary:
            return from_salary
        elif to_salary:
            return to_salary
        else:
            return 0

    @classmethod
    def validate_and_create(cls, data: Dict[str, Any]) -> "Vacancy":
        """Валидирует данные и создает экземпляр Vacancy."""
        salary = data.get("salary")
        if salary is not None:
            if not isinstance(salary, dict):
                raise ValueError("Salary must be a dictionary or None")

        return cls(
            title=data.get("name", ""),
            link=data.get("alternate_url", ""),
            salary=salary,
            description=data.get("description", ""),
            requirements=data.get("snippet", {}).get("requirement", ""),
        )


class VacancyStorage(abc.ABC):
    """Абстрактный класс для работы с хранилищем вакансий."""

    @abc.abstractmethod
    def add_vacancy(self, vacancy: Vacancy) -> None:
        pass

    @abc.abstractmethod
    def get_vacancies(self, criteria: Dict[str, Any]) -> List[Vacancy]:
        pass

    @abc.abstractmethod
    def delete_vacancy(self, criteria: Dict[str, Any]) -> None:
        pass


class JSONVacancyStorage(VacancyStorage):
    """Класс для сохранения вакансий в JSON-файл."""

    def __init__(self, file_path: str):
        self.file_path = file_path

    def add_vacancy(self, vacancy: Vacancy) -> None:
        vacancies = self._load_vacancies()
        # Убеждаемся, что vacancies - это список
        if not isinstance(vacancies, list):
            vacancies = []
        vacancies.append(vacancy.__dict__)
        self._save_vacancies(vacancies)

    def get_vacancies(self, criteria: Dict[str, Any]) -> List[Vacancy]:
        vacancies_data = self._load_vacancies()
        # Убеждаемся, что vacancies_data - это список
        if not isinstance(vacancies_data, list):
            vacancies_data = []
        vacancies = [Vacancy.validate_and_create(data) for data in vacancies_data]
        return self._filter_vacancies(vacancies, criteria)

    def delete_vacancy(self, criteria: Dict[str, Any]) -> None:
        vacancies_data = self._load_vacancies()
        # Убеждаемся, что vacancies_data - это список
        if not isinstance(vacancies_data, list):
            vacancies_data = []
        vacancies = [Vacancy.validate_and_create(data) for data in vacancies_data]
        filtered_vacancies = [
            v for v in vacancies if not self._matches_criteria(v.__dict__, criteria)
        ]
        self._save_vacancies([v.__dict__ for v in filtered_vacancies])

    def _load_vacancies(self) -> List[Dict[str, Any]]:
        try:
            with open(self.file_path, "r", encoding="utf-8") as file:
                data = json.load(file)
                # Проверяем, что загруженные данные - это список
                if isinstance(data, list):
                    return data
                else:
                    print(f"Предупреждение: файл {self.file_path} содержит некорректные данные. Создается новый файл.")
                    return []
        except (FileNotFoundError, json.JSONDecodeError) as e:
            print(f"Файл {self.file_path} не найден или поврежден. Создается новый файл.")
            return []

    def _save_vacancies(self, vacancies: List[Dict[str, Any]]) -> None:
        try:
            # Создаем директорию, если она не существует
            import os
            os.makedirs(os.path.dirname(self.file_path), exist_ok=True)

            with open(self.file_path, "w", encoding="utf-8") as file:
                json.dump(vacancies, file, ensure_ascii=False, indent=4)
        except Exception as e:
            print(f"Ошибка при сохранении файла {self.file_path}: {e}")

    def _filter_vacancies(
            self, vacancies: List[Vacancy], criteria: Dict[str, Any]
    ) -> List[Vacancy]:
        filtered = []
        for vacancy in vacancies:
            matches = True
            for key, value in criteria.items():
                if key == "keyword":
                    if value.lower() not in vacancy.description.lower() and value.lower() not in vacancy.requirements.lower():
                        matches = False
                        break
                elif key == "min_salary":
                    if vacancy.get_salary() < value:
                        matches = False
                        break
                elif getattr(vacancy, key, None) != value:
                    matches = False
                    break
            if matches:
                filtered.append(vacancy)
        return filtered

    def _matches_criteria(self, vacancy_data: Dict[str, Any], criteria: Dict[str, Any]) -> bool:
        for key, value in criteria.items():
            if key == "keyword":
                if value.lower() not in vacancy_data.get("description",
                                                         "").lower() and value.lower() not in vacancy_data.get(
                        "requirements", "").lower():
                    return False
            elif key == "min_salary":
                vacancy = Vacancy.validate_and_create(vacancy_data)
                if vacancy.get_salary() < value:
                    return False
            elif vacancy_data.get(key) != value:
                return False
        return True


class CSVVacancyStorage(VacancyStorage):
    """Класс для сохранения вакансий в CSV-файл."""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self._ensure_file_exists()

    def _ensure_file_exists(self) -> None:
        try:
            # Создаем директорию, если она не существует
            import os
            os.makedirs(os.path.dirname(self.file_path), exist_ok=True)

            with open(self.file_path, "r", newline="", encoding="utf-8") as file:
                pass
        except FileNotFoundError:
            with open(self.file_path, "w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(["title", "link", "salary", "description", "requirements"])
        except Exception as e:
            print(f"Ошибка при создании файла {self.file_path}: {e}")

    def add_vacancy(self, vacancy: Vacancy) -> None:
        with open(self.file_path, "a", newline="", encoding="utf-8") as file:
            writer = csv.writer(file)
            writer.writerow([
                vacancy.title,
                vacancy.link,
                json.dumps(vacancy.salary) if vacancy.salary else "",
                vacancy.description,
                vacancy.requirements,
            ])

    def get_vacancies(self, criteria: Dict[str, Any]) -> List[Vacancy]:
        vacancies = []
        with open(self.file_path, "r", newline="", encoding="utf-8") as file:
            reader = csv.DictReader(file)
            for row in reader:
                salary = json.loads(row["salary"]) if row["salary"] else None
                vacancy = Vacancy(
                    title=row["title"],
                    link=row["link"],
                    salary=salary,
                    description=row["description"],
                    requirements=row["requirements"],
                )
                vacancies.append(vacancy)
        return self._filter_vacancies(vacancies, criteria)

    def delete_vacancy(self, criteria: Dict[str, Any]) -> None:
        vacancies = self.get_vacancies({})
        filtered_vacancies = [v for v in vacancies if not self._matches_criteria(v, criteria)]
        self._save_all_vacancies(filtered_vacancies)

    def _save_all_vacancies(self, vacancies: List[Vacancy]) -> None:
        with open(self.file_path, "w", newline="", encoding="utf-8") as file:
            writer = csv.writer(file)
            writer.writerow(["title", "link", "salary", "description", "requirements"])
            for vacancy in vacancies:
                writer.writerow([
                    vacancy.title,
                    vacancy.link,
                    json.dumps(vacancy.salary) if vacancy.salary else "",
                    vacancy.description,
                    vacancy.requirements,
                ])

    def _filter_vacancies(
        self, vacancies: List[Vacancy], criteria: Dict[str, Any]
    ) -> List[Vacancy]:
        filtered = []
        for vacancy in vacancies:
            matches = True
            for key, value in criteria.items():
                if key == "keyword":
                    if value.lower() not in vacancy.description.lower() and value.lower() not in vacancy.requirements.lower():
                        matches = False
                        break
                elif key == "min_salary":
                    if vacancy.get_salary() < value:
                        matches = False
                        break
                elif getattr(vacancy, key, None) != value:
                    matches = False
                    break
            if matches:
                filtered.append(vacancy)
        return filtered

    def _matches_criteria(self, vacancy: Vacancy, criteria: Dict[str, Any]) -> bool:
        for key, value in criteria.items():
            if key == "keyword":
                if value.lower() not in vacancy.description.lower() and value.lower() not in vacancy.requirements.lower():
                    return False
            elif key == "min_salary":
                if vacancy.get_salary() < value:
                    return False
            elif getattr(vacancy, key, None) != value:
                return False
        return True


class ExcelVacancyStorage(VacancyStorage):
    """Класс для сохранения вакансий в Excel-файл."""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self._ensure_file_exists()

    def _ensure_file_exists(self) -> None:
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            workbook.close()
        except FileNotFoundError:
            # Создаем директорию, если она не существует
            import os
            os.makedirs(os.path.dirname(self.file_path), exist_ok=True)

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["title", "link", "salary", "description", "requirements"])
            workbook.save(self.file_path)
            workbook.close()
        except Exception as e:
            print(f"Ошибка при создании файла {self.file_path}: {e}")

    def add_vacancy(self, vacancy: Vacancy) -> None:
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook.active
        sheet.append([
            vacancy.title,
            vacancy.link,
            json.dumps(vacancy.salary) if vacancy.salary else "",
            vacancy.description,
            vacancy.requirements,
        ])
        workbook.save(self.file_path)
        workbook.close()

    def get_vacancies(self, criteria: Dict[str, Any]) -> List[Vacancy]:
        vacancies = []
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            salary = json.loads(row[2]) if row[2] else None
            vacancy = Vacancy(
                title=row[0],
                link=row[1],
                salary=salary,
                description=row[3],
                requirements=row[4],
            )
            vacancies.append(vacancy)
        workbook.close()
        return self._filter_vacancies(vacancies, criteria)

    def delete_vacancy(self, criteria: Dict[str, Any]) -> None:
        vacancies = self.get_vacancies({})
        filtered_vacancies = [v for v in vacancies if not self._matches_criteria(v, criteria)]
        self._save_all_vacancies(filtered_vacancies)

    def _save_all_vacancies(self, vacancies: List[Vacancy]) -> None:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["title", "link", "salary", "description", "requirements"])
        for vacancy in vacancies:
            sheet.append([
                vacancy.title,
                vacancy.link,
                json.dumps(vacancy.salary) if vacancy.salary else "",
                vacancy.description,
                vacancy.requirements,
            ])
        workbook.save(self.file_path)
        workbook.close()

    def _filter_vacancies(
        self, vacancies: List[Vacancy], criteria: Dict[str, Any]
    ) -> List[Vacancy]:
        filtered = []
        for vacancy in vacancies:
            matches = True
            for key, value in criteria.items():
                if key == "keyword":
                    if value.lower() not in vacancy.description.lower() and value.lower() not in vacancy.requirements.lower():
                        matches = False
                        break
                elif key == "min_salary":
                    if vacancy.get_salary() < value:
                        matches = False
                        break
                elif getattr(vacancy, key, None) != value:
                    matches = False
                    break
            if matches:
                filtered.append(vacancy)
        return filtered

    def _matches_criteria(self, vacancy: Vacancy, criteria: Dict[str, Any]) -> bool:
        for key, value in criteria.items():
            if key == "keyword":
                if value.lower() not in vacancy.description.lower() and value.lower() not in vacancy.requirements.lower():
                    return False
            elif key == "min_salary":
                if vacancy.get_salary() < value:
                    return False
            elif getattr(vacancy, key, None) != value:
                return False
        return True


class TXTVacancyStorage(VacancyStorage):
    """Класс для сохранения вакансий в TXT-файл."""

    def __init__(self, file_path: str):
        self.file_path = file_path

    def add_vacancy(self, vacancy: Vacancy) -> None:
        with open(self.file_path, "a", encoding="utf-8") as file:
            file.write(
                f"{vacancy.title}\t{vacancy.link}\t"
                f"{json.dumps(vacancy.salary) if vacancy.salary else ''}\t"
                f"{vacancy.description}\t{vacancy.requirements}\n"
            )

    def get_vacancies(self, criteria: Dict[str, Any]) -> List[Vacancy]:
        vacancies = []
        try:
            with open(self.file_path, "r", encoding="utf-8") as file:
                for line in file:
                    parts = line.strip().split("\t")
                    if len(parts) != 5:
                        continue
                    title, link, salary_str, description, requirements = parts
                    salary = json.loads(salary_str) if salary_str else None
                    vacancy = Vacancy(
                        title=title,
                        link=link,
                        salary=salary,
                        description=description,
                        requirements=requirements,
                    )
                    vacancies.append(vacancy)
        except FileNotFoundError:
            pass
        return self._filter_vacancies(vacancies, criteria)

    def delete_vacancy(self, criteria: Dict[str, Any]) -> None:
        vacancies = self.get_vacancies({})
        filtered_vacancies = [v for v in vacancies if not self._matches_criteria(v, criteria)]
        self._save_all_vacancies(filtered_vacancies)

    def _save_all_vacancies(self, vacancies: List[Vacancy]) -> None:
        with open(self.file_path, "w", encoding="utf-8") as file:
            for vacancy in vacancies:
                file.write(
                    f"{vacancy.title}\t{vacancy.link}\t"
                    f"{json.dumps(vacancy.salary) if vacancy.salary else ''}\t"
                    f"{vacancy.description}\t{vacancy.requirements}\n"
                )

    def _filter_vacancies(
        self, vacancies: List[Vacancy], criteria: Dict[str, Any]
    ) -> List[Vacancy]:
        filtered = []
        for vacancy in vacancies:
            matches = True
            for key, value in criteria.items():
                if key == "keyword":
                    if value.lower() not in vacancy.description.lower() and value.lower() not in vacancy.requirements.lower():
                        matches = False
                        break
                elif key == "min_salary":
                    if vacancy.get_salary() < value:
                        matches = False
                        break
                elif getattr(vacancy, key, None) != value:
                    matches = False
                    break
            if matches:
                filtered.append(vacancy)
        return filtered

    def _matches_criteria(self, vacancy: Vacancy, criteria: Dict[str, Any]) -> bool:
        for key, value in criteria.items():
            if key == "keyword":
                if value.lower() not in vacancy.description.lower() and value.lower() not in vacancy.requirements.lower():
                    return False
            elif key == "min_salary":
                if vacancy.get_salary() < value:
                    return False
            elif getattr(vacancy, key, None) != value:
                return False
        return True


class VacancyManager:
    """Класс для управления вакансиями."""

    def __init__(self, api: VacancyAPI, storage: VacancyStorage):
        self.api = api
        self.storage = storage

    def fetch_and_store_vacancies(self, search_query: str) -> None:
        """Получает вакансии по API и сохраняет их в хранилище."""
        vacancies_data = self.api.get_vacancies(search_query)
        for data in vacancies_data:
            try:
                vacancy = Vacancy.validate_and_create(data)
                self.storage.add_vacancy(vacancy)
            except ValueError as e:
                print(f"Ошибка при создании вакансии: {e}")

    def get_top_vacancies_by_salary(self, n: int) -> List[Vacancy]:
        """Возвращает топ N вакансий по зарплате."""
        vacancies = self.storage.get_vacancies({})
        vacancies.sort(reverse=True)
        return vacancies[:n]

    def get_vacancies_with_keyword(self, keyword: str) -> List[Vacancy]:
        """Возвращает вакансии, содержащие ключевое слово в описании."""
        return self.storage.get_vacancies({"keyword": keyword})


